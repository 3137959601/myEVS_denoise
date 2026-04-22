from __future__ import annotations

import argparse
import csv
import os
import re
from datetime import datetime
from dataclasses import dataclass
from typing import Any


TEMPLATE_DEFAULT_HEADER = [
    "数据集",
    "噪声程度",
    "算法",
    "s",
    "tau(us)",
    "tpr",
    "fpr",
    "precision",
    "accuracy",
    "f1",
    "auc",
    "esr_mean",
]


EBF_THR_COL = "EBF打分阈值"


@dataclass(frozen=True)
class BestRow:
    dataset: str
    env: str
    method: str
    s: int | None
    tau_us: int | float | None
    tpr: float | None
    fpr: float | None
    precision: float | None
    accuracy: float | None
    f1: float | None
    auc: float | None
    esr_mean: float | None
    # extra
    param: str | None
    value: float | None
    tag: str | None
    source_csv: str


def _to_float(x: Any) -> float | None:
    if x is None:
        return None
    s = str(x).strip()
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None


def _to_int_if_close(x: float | None) -> int | float | None:
    if x is None:
        return None
    xi = int(round(x))
    if abs(x - xi) < 1e-9:
        return xi
    return x


_RE_S = re.compile(r"(?:^|_)s(\d+)(?:$|_)")
_RE_S_TAU = re.compile(r"_s(\d+)_tau(\d+)(?:$|_)")


def _parse_s_from_tag(tag: str | None) -> int | None:
    if not tag:
        return None
    m = _RE_S.search(tag)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


def _parse_s_tau_from_tag(tag: str | None) -> tuple[int | None, int | None]:
    if not tag:
        return None, None
    m = _RE_S_TAU.search(tag)
    if not m:
        return _parse_s_from_tag(tag), None
    try:
        return int(m.group(1)), int(m.group(2))
    except Exception:
        return None, None


def _read_template_header(xlsx_path: str) -> list[str]:
    try:
        import openpyxl  # type: ignore

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb[wb.sheetnames[0]]
        row1 = [c.value for c in ws[1]]
        header = [str(v).strip() for v in row1 if v is not None and str(v).strip()]
        if header:
            return header
    except Exception:
        pass

    return list(TEMPLATE_DEFAULT_HEADER)


def _pick_best_row(csv_path: str, *, method: str, env: str, dataset: str) -> BestRow:
    if not os.path.exists(csv_path):
        raise FileNotFoundError(csv_path)

    rows: list[dict[str, str]] = []
    with open(csv_path, "r", newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        for row in r:
            rows.append({k: (v or "") for k, v in row.items() if k is not None})

    if not rows:
        raise RuntimeError(f"empty roc csv: {csv_path}")

    # Selection rule (configurable via env for backward-compat):
    # - primary=AUC: maximize AUC first, then maximize F1 under that AUC.
    # - primary=F1:  maximize F1 first, then maximize AUC under that F1.
    primary = (os.environ.get("MYEVS_SUMMARY_PRIMARY", "auc") or "auc").strip().lower()
    if primary not in {"auc", "f1"}:
        primary = "auc"

    def _row_key_auc_then_f1(row: dict[str, str]) -> tuple[float, float, float]:
        auc = _to_float(row.get("auc"))
        f1 = _to_float(row.get("f1"))
        tpr = _to_float(row.get("tpr"))
        aucv = float(auc) if auc is not None else -1.0
        f1v = float(f1) if f1 is not None else -1.0
        tprv = float(tpr) if tpr is not None else 0.0
        return (aucv, f1v, tprv)

    def _row_key_f1_then_auc(row: dict[str, str]) -> tuple[float, float, float]:
        f1 = _to_float(row.get("f1"))
        auc = _to_float(row.get("auc"))
        tpr = _to_float(row.get("tpr"))
        f1v = float(f1) if f1 is not None else -1.0
        aucv = float(auc) if auc is not None else -1.0
        tprv = float(tpr) if tpr is not None else 0.0
        return (f1v, aucv, tprv)

    best = max(rows, key=_row_key_auc_then_f1 if primary == "auc" else _row_key_f1_then_auc)

    tag = (best.get("tag") or "").strip() or None
    param = (best.get("param") or "").strip() or None
    value = _to_float(best.get("value"))

    # Extract s/tau.
    s: int | None
    tau_us: int | float | None

    if method.upper() == "EBF":
        s_tag, tau_tag = _parse_s_tau_from_tag(tag)
        s = s_tag
        tau_us = tau_tag
    else:
        s = _parse_s_from_tag(tag)
        # For BAF/FDF/STCF, tau is the swept parameter value (time-us)
        tau_us = _to_int_if_close(_to_float(best.get("value")))

    return BestRow(
        dataset=dataset,
        env=env,
        method=method,
        s=s,
        tau_us=tau_us,
        tpr=_to_float(best.get("tpr")),
        fpr=_to_float(best.get("fpr")),
        precision=_to_float(best.get("precision")),
        accuracy=_to_float(best.get("accuracy")),
        f1=_to_float(best.get("f1")),
        auc=_to_float(best.get("auc")),
        esr_mean=_to_float(best.get("esr_mean")),
        param=param,
        value=value,
        tag=tag,
        source_csv=csv_path.replace("\\", "/"),
    )


def main() -> int:
    ap = argparse.ArgumentParser(
        description=(
            "Summarize best params for 4 methods (BAF/EBF/FDF/STCF) on ED24/myPedestrain_06. "
            "Selection rule: maximize AUC first, then maximize F1 under that AUC."
        )
    )
    ap.add_argument(
        "--root",
        default="data/ED24/myPedestrain_06",
        help="Root directory containing BAF/EBF/FDF/STCF subfolders",
    )
    ap.add_argument(
        "--dataset",
        default="myPedestrain_06",
        help="Value for the '数据集' column",
    )
    ap.add_argument(
        "--template-xlsx",
        default="data/算法数据汇总表.xlsx",
        help="Excel template to reference header names",
    )
    ap.add_argument(
        "--out-csv",
        default="data/ED24/myPedestrain_06/算法最优参数汇总_auc优先_f1次之.csv",
        help="Output CSV path",
    )
    ap.add_argument(
        "--methods",
        default="BAF,FDF,STCF,EBF",
        help="Comma-separated subset of methods to summarize, e.g. 'EBF' or 'BAF,EBF'",
    )
    ap.add_argument(
        "--primary-metric",
        choices=["auc", "f1"],
        default="auc",
        help=(
            "Primary metric to pick best operating point. "
            "auc: maximize AUC first, then F1; f1: maximize F1 first, then AUC."
        ),
    )
    ap.add_argument(
        "--ebf-roc-prefix",
        default="roc_ebf",
        help=(
            "ROC CSV filename prefix for EBF family. "
            "Use 'roc_ebf' for baseline, 'roc_ebf_v10' for V10 outputs."
        ),
    )
    ap.add_argument(
        "--ebf-dir",
        default="",
        help=(
            "Optional directory containing EBF-family ROC CSVs. "
            "If empty, defaults to <root>/EBF. "
            "Useful for Part2 experiments stored under EBF_Part2 subfolders."
        ),
    )
    ap.add_argument(
        "--ebf-method-name",
        default="EBF",
        help="Method name written into the output CSV for EBF rows (e.g. EBF / EBFV10).",
    )
    args = ap.parse_args()

    root = str(args.root)
    dataset = str(args.dataset)
    # propagate selection rule to _pick_best_row (keeps existing signature)
    os.environ["MYEVS_SUMMARY_PRIMARY"] = str(args.primary_metric)

    # File naming uses light/mid/heavy, but we display 1.8V/2.5V/3.3V in the summary.
    env_keys = ["light", "mid", "heavy"]
    env_label = {"light": "1.8V", "mid": "2.5V", "heavy": "3.3V"}
    # Keep the output order consistent with the user's spreadsheet convention.
    wanted = [m.strip().upper() for m in str(args.methods).split(",") if m.strip()]
    methods_all = ["BAF", "FDF", "STCF", "EBF"]
    methods = [m for m in methods_all if m in set(wanted)]
    if not methods:
        raise SystemExit(f"--methods empty/invalid: {args.methods!r}. choices={methods_all}")

    # Locate ROC CSV for each method/env.
    def roc_csv_path(method: str, env: str) -> str:
        if method == "BAF":
            return os.path.join(root, "BAF", f"roc_baf_{env}.csv")
        if method == "FDF":
            return os.path.join(root, "FDF", f"roc_fdf_{env}.csv")
        if method == "STCF":
            return os.path.join(root, "STCF", f"roc_stcf_{env}.csv")
        if method == "EBF":
            ebf_dir = (str(args.ebf_dir).strip() or "").replace("\\", "/")
            if not ebf_dir:
                ebf_dir = os.path.join(root, "EBF")
            # Prefer labelscore files (from sweep_ebf_labelscore_grid.py)
            prefix = str(args.ebf_roc_prefix).strip() or "roc_ebf"
            patt = re.compile(rf"^{re.escape(prefix)}_{re.escape(env)}_labelscore.*\.csv$", re.IGNORECASE)
            cand = [
                os.path.join(ebf_dir, fn)
                for fn in os.listdir(ebf_dir)
                if os.path.isfile(os.path.join(ebf_dir, fn)) and patt.match(fn)
            ]
            if cand:
                return max(cand, key=lambda p: os.path.getmtime(p))
            # Fallbacks
            p2 = os.path.join(ebf_dir, f"{prefix}_{env}_label_exact.csv")
            if os.path.exists(p2):
                return p2
            p3 = os.path.join(ebf_dir, f"{prefix}_{env}.csv")
            return p3
        raise ValueError(f"unknown method: {method}")

    best_rows: list[BestRow] = []
    for env in env_keys:
        for method in methods:
            p = roc_csv_path(method, env)
            best_rows.append(_pick_best_row(p, method=method, env=env, dataset=dataset))

    # Header: follow the Excel template, and insert an extra EBF threshold column.
    header = _read_template_header(str(args.template_xlsx))
    header = [h for h in header if h in TEMPLATE_DEFAULT_HEADER] or list(TEMPLATE_DEFAULT_HEADER)
    # Ensure ESR column exists even if the Excel template doesn't have it.
    if "esr_mean" not in header:
        if "auc" in header:
            i_auc = header.index("auc") + 1
            header = header[:i_auc] + ["esr_mean"] + header[i_auc:]
        else:
            header = header + ["esr_mean"]
    if EBF_THR_COL not in header:
        # Put the extra column right after "算法".
        if "算法" in header:
            i = header.index("算法") + 1
            out_header = header[:i] + [EBF_THR_COL] + header[i:]
        else:
            out_header = header + [EBF_THR_COL]
    else:
        out_header = header

    out_csv = str(args.out_csv)
    os.makedirs(os.path.dirname(os.path.abspath(out_csv)), exist_ok=True)
    # Use UTF-8 with BOM so Excel on Windows opens Chinese headers correctly.
    def _write_csv(path: str) -> None:
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=out_header)
            w.writeheader()
            for br in best_rows:
                ebf_thr = ""
                if br.method.upper() == "EBF" and br.value is not None:
                    ebf_thr = br.value
                method_name = br.method
                if br.method.upper() == "EBF":
                    method_name = str(args.ebf_method_name).strip() or "EBF"
                w.writerow(
                    {
                        "数据集": br.dataset,
                        "噪声程度": env_label.get(br.env, br.env),
                        "算法": method_name,
                        EBF_THR_COL: ebf_thr,
                        "s": br.s if br.s is not None else "",
                        "tau(us)": br.tau_us if br.tau_us is not None else "",
                        "tpr": br.tpr if br.tpr is not None else "",
                        "fpr": br.fpr if br.fpr is not None else "",
                        "precision": br.precision if br.precision is not None else "",
                        "accuracy": br.accuracy if br.accuracy is not None else "",
                        "f1": br.f1 if br.f1 is not None else "",
                        "auc": br.auc if br.auc is not None else "",
                        "esr_mean": br.esr_mean if br.esr_mean is not None else "",
                    }
                )

    try:
        _write_csv(out_csv)
    except PermissionError:
        base, ext = os.path.splitext(out_csv)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_csv = f"{base}_{ts}{ext or '.csv'}"
        _write_csv(out_csv)

    print(f"saved: {out_csv}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
