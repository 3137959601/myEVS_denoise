from __future__ import annotations

import argparse
import csv
import os
import re
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Iterable

import math


_ENV_KEYS = ["light", "mid", "heavy"]
_ENV_LABEL = {"light": "1.8V", "mid": "2.5V", "heavy": "3.3V"}


TEMPLATE_DEFAULT_HEADER = [
    "数据集",
    "噪声程度",
    "算法",
    "s",
    "tau(us)",
    "tpr",
    "fpr",
    "precision",
    "accuracy",
    "f1",
    "auc",
    "esr_mean",
]


THR_COL = "EBF_optimized打分阈值"

_RE_S_TAU = re.compile(r"_s(\d+)_tau(\d+)(?:$|_)", re.IGNORECASE)


@dataclass(frozen=True)
class BestTagResult:
    env: str
    dataset: str
    tag: str
    s: int | None
    tau_us: int | None
    thr: float | None
    f1: float | None
    tpr: float | None
    fpr: float | None
    precision: float | None
    accuracy: float | None
    auc: float | None
    esr_mean: float | None
    source_csv: str


def _to_float(x: Any) -> float | None:
    if x is None:
        return None
    s = str(x).strip()
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None


def _parse_s_tau(tag: str) -> tuple[int | None, int | None]:
    m = _RE_S_TAU.search(tag or "")
    if not m:
        return None, None
    try:
        return int(m.group(1)), int(m.group(2))
    except Exception:
        return None, None


def _read_rows(csv_path: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    with open(csv_path, "r", newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        for row in r:
            if not row:
                continue
            rows.append({k: (v or "") for k, v in row.items() if k is not None})
    return rows


def _group_by_tag(rows: Iterable[dict[str, str]]) -> dict[str, list[dict[str, str]]]:
    g: dict[str, list[dict[str, str]]] = {}
    for row in rows:
        tag = (row.get("tag") or "").strip()
        if not tag:
            continue
        g.setdefault(tag, []).append(row)
    return g


def _pick_best_row_in_rows(rows: list[dict[str, str]]) -> dict[str, str]:
    """Pick best operating point (threshold) within a list of ROC rows.

    Primary metric: max F1
    Tie-breakers: higher TPR, then higher precision, then lower FPR.

    Returns one row dict.
    """

    def key(row: dict[str, str]) -> tuple[float, float, float, float]:
        f1 = _to_float(row.get("f1"))
        tpr = _to_float(row.get("tpr"))
        prec = _to_float(row.get("precision"))
        fpr = _to_float(row.get("fpr"))

        f1v = float(f1) if f1 is not None else -1.0
        tprv = float(tpr) if tpr is not None else 0.0
        precv = float(prec) if prec is not None else 0.0
        fprv = float(fpr) if fpr is not None else 1.0
        return (f1v, tprv, precv, -fprv)

    return max(rows, key=key)


def _pick_best_tag(rows: list[dict[str, str]]) -> tuple[str, float] | None:
    """Pick best tag (s,tau) primarily by max AUC; tie by best-F1 of that tag."""

    by_tag = _group_by_tag(rows)
    if not by_tag:
        return None

    best_tag: str = ""
    best_auc: float = -1.0
    best_tag_best_f1: float = -1.0

    for tag, tag_rows in by_tag.items():
        # AUC is identical for all rows in a tag; take first valid
        auc_val: float | None = None
        for r in tag_rows:
            auc_val = _to_float(r.get("auc"))
            if auc_val is not None:
                break
        if auc_val is None:
            continue

        best_row = _pick_best_row_in_rows(tag_rows)
        f1_val = _to_float(best_row.get("f1"))
        f1v = float(f1_val) if f1_val is not None else -1.0

        if auc_val > best_auc + 1e-15:
            best_tag = tag
            best_auc = float(auc_val)
            best_tag_best_f1 = f1v
        elif abs(auc_val - best_auc) <= 1e-15:
            if f1v > best_tag_best_f1 + 1e-15:
                best_tag = tag
                best_auc = float(auc_val)
                best_tag_best_f1 = f1v

    if not best_tag:
        return None
    return best_tag, best_auc


def _ensure_dir(path: str) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)


def _write_csv(path: str, header: list[str], rows: list[dict[str, Any]]) -> str:
    _ensure_dir(path)
    try:
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=header)
            w.writeheader()
            for r in rows:
                w.writerow(r)
        return path
    except PermissionError:
        base, ext = os.path.splitext(path)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = f"{base}_{ts}{ext or '.csv'}"
        with open(alt, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=header)
            w.writeheader()
            for r in rows:
                w.writerow(r)
        return alt


def _find_default_env_csv(in_dir: str, env: str) -> str:
    # Compatible naming:
    # - legacy: roc_ebf_optimized_{env}_labelscore_*.csv
    # - new:    roc_ebf_optimized_{variant}_{env}_labelscore_*.csv
    patt = re.compile(
        rf"^roc_ebf_optimized_(?:[a-z0-9_]+_)?{re.escape(env)}_labelscore.*\.csv$",
        re.IGNORECASE,
    )
    if not os.path.isdir(in_dir):
        raise FileNotFoundError(in_dir)
    cand = [
        os.path.join(in_dir, fn)
        for fn in os.listdir(in_dir)
        if os.path.isfile(os.path.join(in_dir, fn)) and patt.match(fn)
    ]
    if not cand:
        raise FileNotFoundError(f"missing roc csv for env={env} under: {in_dir}")
    return max(cand, key=lambda p: os.path.getmtime(p))


def summarize_one_env(
    *,
    csv_path: str,
    env: str,
    dataset: str,
    force_tag: str | None,
) -> BestTagResult:
    rows = _read_rows(csv_path)
    if not rows:
        raise RuntimeError(f"empty roc csv: {csv_path}")

    if force_tag:
        tag = force_tag
    else:
        best = _pick_best_tag(rows)
        if best is None:
            raise RuntimeError(f"no valid tag/auc in: {csv_path}")
        tag = best[0]

    tag_rows = [r for r in rows if (r.get("tag") or "").strip() == tag]
    if not tag_rows:
        raise RuntimeError(f"tag not found: {tag} in {csv_path}")

    best_row = _pick_best_row_in_rows(tag_rows)

    s, tau_us = _parse_s_tau(tag)
    return BestTagResult(
        env=env,
        dataset=dataset,
        tag=tag,
        s=s,
        tau_us=tau_us,
        thr=_to_float(best_row.get("value")),
        f1=_to_float(best_row.get("f1")),
        tpr=_to_float(best_row.get("tpr")),
        fpr=_to_float(best_row.get("fpr")),
        precision=_to_float(best_row.get("precision")),
        accuracy=_to_float(best_row.get("accuracy")),
        auc=_to_float(best_row.get("auc")),
        esr_mean=_to_float(best_row.get("esr_mean")),
        source_csv=csv_path.replace("\\\\", "/"),
    )


def _thr_stats(thrs: list[float]) -> dict[str, float]:
    if not thrs:
        return {"min": math.nan, "max": math.nan, "range": math.nan, "mean": math.nan, "std": math.nan}
    mn = min(thrs)
    mx = max(thrs)
    mean = sum(thrs) / float(len(thrs))
    var = sum((x - mean) ** 2 for x in thrs) / float(len(thrs))
    std = math.sqrt(var)
    return {"min": mn, "max": mx, "range": mx - mn, "mean": mean, "std": std}


def _read_template_header(xlsx_path: str) -> list[str]:
    try:
        import openpyxl  # type: ignore

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb[wb.sheetnames[0]]
        row1 = [c.value for c in ws[1]]
        header = [str(v).strip() for v in row1 if v is not None and str(v).strip()]
        if header:
            return header
    except Exception:
        pass

    return list(TEMPLATE_DEFAULT_HEADER)


def main() -> int:
    ap = argparse.ArgumentParser(
        description=(
            "Summarize best params for EBF_optimized (s,tau,Thr) from ROC CSV and report Thr stability across envs. "
            "Best tag is chosen by AUC; best Thr within that tag is chosen by max F1."
        )
    )
    ap.add_argument(
        "--in-dir",
        default="data/ED24/myPedestrain_06/EBF_optimized",
        help="Directory containing roc_ebf_optimized_{light,mid,heavy}_*.csv",
    )
    ap.add_argument("--dataset", default="myPedestrain_06")
    ap.add_argument(
        "--template-xlsx",
        default="data/算法数据汇总表.xlsx",
        help="Excel template to reference header names (optional)",
    )
    ap.add_argument(
        "--force-tag",
        default="",
        help=(
            "If set, do NOT search best tag; instead evaluate this tag in all envs and pick best Thr by F1. "
            "Example: ebfopt_labelscore_s9_tau128000"
        ),
    )
    ap.add_argument(
        "--out-csv",
        default="data/ED24/myPedestrain_06/EBF_optimized/best_params_ebf_optimized.csv",
        help="Per-env summary output (CSV)",
    )
    ap.add_argument(
        "--out-stability-csv",
        default="data/ED24/myPedestrain_06/EBF_optimized/thr_stability_ebf_optimized.csv",
        help="Thr stability output (CSV)",
    )
    args = ap.parse_args()

    in_dir = str(args.in_dir)
    dataset = str(args.dataset)
    force_tag = str(args.force_tag).strip() or None

    env_csv = {env: _find_default_env_csv(in_dir, env) for env in _ENV_KEYS}

    results: list[BestTagResult] = []
    for env in _ENV_KEYS:
        results.append(
            summarize_one_env(
                csv_path=env_csv[env],
                env=env,
                dataset=dataset,
                force_tag=force_tag,
            )
        )

    # Determine whether tags are consistent across envs
    tags = [r.tag for r in results]
    tag_unique = sorted(set(tags))

    # Per-env output (template-like + extra debug cols)
    header_base = _read_template_header(str(args.template_xlsx))
    header_base = [h for h in header_base if h in TEMPLATE_DEFAULT_HEADER] or list(TEMPLATE_DEFAULT_HEADER)
    # Ensure ESR column exists even if the Excel template doesn't have it.
    if "esr_mean" not in header_base:
        if "auc" in header_base:
            i_auc = header_base.index("auc") + 1
            header_base = header_base[:i_auc] + ["esr_mean"] + header_base[i_auc:]
        else:
            header_base = header_base + ["esr_mean"]
    if THR_COL not in header_base:
        if "算法" in header_base:
            i = header_base.index("算法") + 1
            out_header = header_base[:i] + [THR_COL] + header_base[i:]
        else:
            out_header = header_base + [THR_COL]
    else:
        out_header = header_base
    # Keep tag/source for traceability (safe to ignore in Excel)
    if "tag" not in out_header:
        out_header = out_header + ["tag"]
    if "source_csv" not in out_header:
        out_header = out_header + ["source_csv"]

    out_rows: list[dict[str, Any]] = []
    for r in results:
        out_rows.append(
            {
                "数据集": r.dataset,
                "噪声程度": _ENV_LABEL.get(r.env, r.env),
                "算法": "EBF_optimized",
                "s": r.s if r.s is not None else "",
                "tau(us)": r.tau_us if r.tau_us is not None else "",
                "f1": r.f1 if r.f1 is not None else "",
                "tpr": r.tpr if r.tpr is not None else "",
                "fpr": r.fpr if r.fpr is not None else "",
                "precision": r.precision if r.precision is not None else "",
                "accuracy": r.accuracy if r.accuracy is not None else "",
                "auc": r.auc if r.auc is not None else "",
                "esr_mean": r.esr_mean if r.esr_mean is not None else "",
                THR_COL: r.thr if r.thr is not None else "",
                "tag": r.tag,
                "source_csv": r.source_csv,
            }
        )

    out_path = _write_csv(str(args.out_csv), out_header, out_rows)

    # Stability output (only meaningful when comparing SAME tag across envs)
    stability_rows: list[dict[str, Any]] = []
    if len(tag_unique) == 1:
        thrs = [float(r.thr) for r in results if r.thr is not None]
        stats = _thr_stats(thrs)
        stability_rows.append(
            {
                "数据集": dataset,
                "算法": "EBF_optimized",
                "tag": tag_unique[0],
                "s": results[0].s if results and results[0].s is not None else "",
                "tau(us)": results[0].tau_us if results and results[0].tau_us is not None else "",
                "Thr_light": next((r.thr for r in results if r.env == "light"), ""),
                "Thr_mid": next((r.thr for r in results if r.env == "mid"), ""),
                "Thr_heavy": next((r.thr for r in results if r.env == "heavy"), ""),
                "Thr_min": stats["min"],
                "Thr_max": stats["max"],
                "Thr_range": stats["range"],
                "Thr_mean": stats["mean"],
                "Thr_std": stats["std"],
            }
        )
    else:
        stability_rows.append(
            {
                "数据集": dataset,
                "算法": "EBF_optimized",
                "tag": "",
                "s": "",
                "tau(us)": "",
                "Thr_light": "",
                "Thr_mid": "",
                "Thr_heavy": "",
                "Thr_min": "",
                "Thr_max": "",
                "Thr_range": "",
                "Thr_mean": "",
                "Thr_std": "",
                "note": "tags differ across envs; use --force-tag to compare a fixed (s,tau)",
            }
        )

    stability_header = [
        "数据集",
        "算法",
        "tag",
        "s",
        "tau(us)",
        "Thr_light",
        "Thr_mid",
        "Thr_heavy",
        "Thr_min",
        "Thr_max",
        "Thr_range",
        "Thr_mean",
        "Thr_std",
    ]
    if "note" in stability_rows[0]:
        stability_header.append("note")

    stability_path = _write_csv(str(args.out_stability_csv), stability_header, stability_rows)

    print(f"saved: {out_path}")
    print(f"saved: {stability_path}")
    print("=== TAGS (by env) ===")
    for r in results:
        print(f"{r.env}: {r.tag} | Thr={r.thr} | F1={r.f1} | AUC={r.auc}")

    if len(tag_unique) == 1:
        stats = stability_rows[0]
        print("=== THR STABILITY (same tag across envs) ===")
        print(
            "tag={tag} | Thr(light,mid,heavy)=({l},{m},{h}) | range={rg} | std={sd}".format(
                tag=stats["tag"],
                l=stats["Thr_light"],
                m=stats["Thr_mid"],
                h=stats["Thr_heavy"],
                rg=stats["Thr_range"],
                sd=stats["Thr_std"],
            )
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
